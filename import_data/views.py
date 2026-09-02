import pandas as pd
from django.shortcuts import render
from django.contrib import messages
from django.db import transaction
from .forms import MultipleImportForm
from Collaborateur.models import Collaborateur, Departement, Unite
from declaration_effectif.models import historique
from utilisateur.decorators import role_required
from declaration_effectif.models import declaration_effectif
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
from django.http import HttpResponse
# ------------------------------------------------------------------
# FONCTIONS AUXILIAIRES
# ------------------------------------------------------------------

def clean_val(val):
    if val is None or pd.isna(val):
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()

def clean_int(val, default=0):
    try:
        if pd.isna(val) or val is None:
            return default
        return int(val)
    except (ValueError, TypeError):
        return default

def read_uploaded_file(fichier):
    if fichier.name.endswith(".csv"):
        df = pd.read_csv(fichier)
    else:
        df = pd.read_excel(fichier)
    df.columns = df.columns.str.strip()
    return df

# ------------------------------------------------------------------
# IMPORT DÉPARTEMENTS
# ------------------------------------------------------------------

def importer_departements(df_dpt, erreurs):
    collaborateurs_map = {c.it: c for c in Collaborateur.objects.all()}
    departements_existants = {d.abreviation: d for d in Departement.objects.all()}

    a_creer = []
    a_maj = []

    for i, row in df_dpt.iterrows():
        try:
            abbrev = clean_val(row.get("Abreviation") or row.get("abreviation"))
            if not abbrev:
                erreurs.append(f"Départements Ligne {i+2}: Abréviation manquante.")
                continue

            nom_dpt = row.get("nom_departement") or abbrev
            rh_val = clean_val(row.get("HRBP"))
            admin_val = clean_val(row.get("ADMIN"))
            drh_val = clean_val(row.get("DRH"))

            rh_obj = collaborateurs_map.get(rh_val) if rh_val else None
            admin_obj = collaborateurs_map.get(admin_val) if admin_val else None
            drh_obj = collaborateurs_map.get(drh_val) if drh_val else None

            if rh_val and not rh_obj:
                erreurs.append(f"Départements Ligne {i+2}: HRBP '{rh_val}' introuvable.")
            if admin_val and not admin_obj:
                erreurs.append(f"Départements Ligne {i+2}: ADMIN '{admin_val}' introuvable.")
            if drh_val and not drh_obj:
                erreurs.append(f"Départements Ligne {i+2}: DRH '{drh_val}' introuvable.")

            data = dict(
                nom_departement=nom_dpt,
                HRBP=rh_obj,
                ADMIN=admin_obj,
                DRH=drh_obj,
                maquette=clean_int(row.get("Maquette")),
            )

            if abbrev in departements_existants:
                obj = departements_existants[abbrev]
                for k, v in data.items():
                    setattr(obj, k, v)
                a_maj.append(obj)
            else:
                obj = Departement(abreviation=abbrev, **data)
                a_creer.append(obj)
                departements_existants[abbrev] = obj

        except Exception as e:
            erreurs.append(f"Départements Ligne {i+2}: {e}")

    with transaction.atomic():
        if a_creer:
            Departement.objects.bulk_create(a_creer, batch_size=500)
        if a_maj:
            Departement.objects.bulk_update(
                a_maj, ["nom_departement", "HRBP", "ADMIN", "DRH", "maquette"], batch_size=500
            )

    return len(a_creer) + len(a_maj)

# ------------------------------------------------------------------
# IMPORT UNITÉS
# ------------------------------------------------------------------

def importer_unites(df_unite, erreurs):
    unites_existantes = {u.abreviation: u for u in Unite.objects.all()}
    a_creer = []
    a_maj = []

    for i, row in df_unite.iterrows():
        try:
            abbrev = clean_val(row.get("abreviation") or row.get("Abreviation"))
            if not abbrev:
                erreurs.append(f"Unités Ligne {i+2}: Abréviation manquante.")
                continue

            data = dict(
                nom=row.get("nom") or abbrev,
                maquette=clean_int(row.get("maquette"), 0),
                A=clean_int(row.get("A"), 0),
                T=clean_int(row.get("T"), 0),
                P=clean_int(row.get("P"), 0),
                C=clean_int(row.get("C"), 0),
            )

            if abbrev in unites_existantes:
                obj = unites_existantes[abbrev]
                for k, v in data.items():
                    setattr(obj, k, v)
                a_maj.append(obj)
            else:
                obj = Unite(abreviation=abbrev, **data)
                a_creer.append(obj)
                unites_existantes[abbrev] = obj

        except Exception as e:
            erreurs.append(f"Unités Ligne {i+2}: {e}")

    with transaction.atomic():
        if a_creer:
            Unite.objects.bulk_create(a_creer, batch_size=500)
        if a_maj:
            Unite.objects.bulk_update(a_maj, ["nom", "maquette", "A", "T", "P", "C"], batch_size=500)

    return len(a_creer) + len(a_maj)

# ------------------------------------------------------------------
# IMPORT COLLABORATEURS
# ------------------------------------------------------------------

def importer_collaborateurs(df_collab, erreurs):
    departements_map = {d.abreviation: d for d in Departement.objects.all()}
    unites_map = {u.abreviation: u for u in Unite.objects.all()}
    collaborateurs_map = {c.it: c for c in Collaborateur.objects.all()}
    collaborateurs_par_matricule = {
        c.matricule: c.it for c in collaborateurs_map.values() if c.matricule
    }

    a_creer = []
    a_maj = []
    ru_a_resoudre = []

    for i, row in df_collab.iterrows():
        try:
            utilisateur_it = clean_val(row.get("Utilisateur"))
            if not utilisateur_it:
                erreurs.append(f"Collaborateurs Ligne {i+2}: Identifiant Utilisateur manquant.")
                continue

            dpt_code = clean_val(row.get("DPT"))
            dpt_obj = departements_map.get(dpt_code) if dpt_code else None
            if dpt_code and not dpt_obj:
                erreurs.append(f"Collaborateurs Ligne {i+2}: Département '{dpt_code}' introuvable.")

            unite_code = clean_val(row.get("Unite"))
            unite_obj = unites_map.get(unite_code) if unite_code else None
            if unite_code and not unite_obj:
                erreurs.append(f"Collaborateurs Ligne {i+2}: Unité '{unite_code}' introuvable.")

            ru_mat = clean_val(row.get("RU"))
            nom = str(row.get("Nom") or "").strip()
            prenom = str(row.get("Prénom") or "").strip()

            data = dict(
                matricule=clean_val(row.get("Matricule")),
                nom_complete=f"{nom} {prenom}".strip(),
                lot=str(row.get("Lot", "") or "").strip(),
                departement=dpt_obj,
                unite=unite_obj,
                eq=str(row.get("Equipe", "") or ""),
                shift=row.get("Shift"),
                sexe=clean_int(row.get("Sexe"), default=1),
            )

            matricule_val = data["matricule"]

            if utilisateur_it in collaborateurs_map:
                obj = collaborateurs_map[utilisateur_it]
                for k, v in data.items():
                    setattr(obj, k, v)
                a_maj.append(obj)
            else:
                obj = Collaborateur(it=utilisateur_it, **data)
                a_creer.append(obj)
                collaborateurs_map[utilisateur_it] = obj

            if matricule_val:
                collaborateurs_par_matricule[matricule_val] = utilisateur_it

            if ru_mat:
                ru_a_resoudre.append((utilisateur_it, ru_mat))

        except Exception as e:
            erreurs.append(f"Collaborateurs Ligne {i+2}: {e}")

    with transaction.atomic():
        if a_creer:
            Collaborateur.objects.bulk_create(a_creer, batch_size=500)
        if a_maj:
            Collaborateur.objects.bulk_update(
                a_maj,
                ["matricule", "nom_complete", "lot", "departement", "unite", "eq", "shift", "sexe"],
                batch_size=500,
            )

    if ru_a_resoudre:
        tous_les_collabs = {
            c.it: c for c in Collaborateur.objects.filter(it__in=[u for u, _ in ru_a_resoudre])
        }

        a_maj_ru = []
        for utilisateur_it, ru_val in ru_a_resoudre:
            collab = tous_les_collabs.get(utilisateur_it)
            if collab is None:
                continue

            ru_it_resolu = ru_val if ru_val in collaborateurs_map else None
            if not ru_it_resolu:
                ru_it_resolu = collaborateurs_par_matricule.get(ru_val)

            if ru_it_resolu:
                collab.ru_it_id = ru_it_resolu
                a_maj_ru.append(collab)
            else:
                erreurs.append(
                    f"RU '{ru_val}' introuvable (ni comme 'it', ni comme matricule) pour le collaborateur '{utilisateur_it}'."
                )

        if a_maj_ru:
            Collaborateur.objects.bulk_update(a_maj_ru, ["ru_it"], batch_size=500)

    return len(a_creer) + len(a_maj)

# ------------------------------------------------------------------
# IMPORT CHANGEMENTS D'AFFECTATION
# ------------------------------------------------------------------

def importer_changements(df_chg, erreurs):
    departements_map = {d.abreviation: d for d in Departement.objects.all()}
    a_creer = []

    for i, row in df_chg.iterrows():
        try:
            collaborateur = clean_val(row.get("Nom & prénom"))
            if not collaborateur:
                erreurs.append(f"Changements Ligne {i+2}: Nom du collaborateur manquant.")
                continue

            initial = clean_val(row.get("Ru (Initial)")) or ""
            acceuil = clean_val(row.get("RU D'accueil")) or ""
            etat = clean_val(row.get("Etat")) or ""

            dpt_init_code = clean_val(row.get("DPT"))
            dpt_acceuil_code = clean_val(row.get("DPT (accueil)"))

            dpt_init_obj = departements_map.get(dpt_init_code) if dpt_init_code else None
            dpt_acceuil_obj = departements_map.get(dpt_acceuil_code) if dpt_acceuil_code else None

            if dpt_init_code and not dpt_init_obj:
                erreurs.append(f"Changements Ligne {i+2}: Département initial '{dpt_init_code}' introuvable.")
            if dpt_acceuil_code and not dpt_acceuil_obj:
                erreurs.append(f"Changements Ligne {i+2}: Département d'accueil '{dpt_acceuil_code}' introuvable.")

            obj = historique(
                collaborateur=collaborateur[:30],
                initial=initial[:30],
                acceuil=acceuil[:30],
                etat=etat[:30],
                dpt_init=dpt_init_obj,
                dpt_acceuil=dpt_acceuil_obj,
            )
            a_creer.append(obj)
        except Exception as e:
            erreurs.append(f"Changements Ligne {i+2}: {e}")

    with transaction.atomic():
        if a_creer:
            historique.objects.bulk_create(a_creer, batch_size=500)

    return len(a_creer)

# ------------------------------------------------------------------
# VUE PRINCIPALE
# ------------------------------------------------------------------

@role_required(['HRBP', 'ADMIN', 'SUPER'])
def importer_fichiers_combines(request):
    role = request.session.get('role')
    template_de_base = {
        "HRBP":  "utilisateur/navbar_N1.html",
        "ADMIN": "utilisateur/navbar_N1.html",
        "SUPER": "utilisateur/navbar_N1.html",
    }.get(role, "utilisateur/navbar_N1.html")

    if request.method != "POST":
        storage = messages.get_messages(request)
        for _ in storage:
            pass
        
        return render(request, "import_data/import.html", {
            "form": MultipleImportForm(),
            "template_de_base": template_de_base,
        })

    form = MultipleImportForm(request.POST, request.FILES)
    if not form.is_valid():
        return render(request, "import_data/import.html", {"form": form, "template_de_base": template_de_base})

    f_collab = request.FILES.get("fichier_collaborateur")
    f_unite = request.FILES.get("fichier_unite")
    f_dpt = request.FILES.get("fichier_departement")
    f_chg = request.FILES.get("fichier_changement")  

    if not (f_collab or f_unite or f_dpt or f_chg):
        messages.error(request, "Veuillez fournir au moins un fichier à importer.")
        return render(request, "import_data/import.html", {"form": form, "template_de_base": template_de_base})

    erreurs = []
    crees_dpts = crees_unites = crees_collabs = crees_chgs = 0

    if f_dpt:
        try:
            df_dpt = read_uploaded_file(f_dpt)
            crees_dpts = importer_departements(df_dpt, erreurs)
        except Exception as e:
            messages.error(request, f"Erreur de lecture du fichier Départements : {e}")

    if f_unite:
        try:
            df_unite = read_uploaded_file(f_unite)
            crees_unites = importer_unites(df_unite, erreurs)
        except Exception as e:
            messages.error(request, f"Erreur de lecture du fichier Unités : {e}")

    if f_collab:
        try:
            df_collab = read_uploaded_file(f_collab)
            crees_collabs = importer_collaborateurs(df_collab, erreurs)
        except Exception as e:
            messages.error(request, f"Erreur de lecture du fichier Collaborateurs : {e}")

    if f_chg:
        try:
            df_chg = read_uploaded_file(f_chg)
            crees_chgs = importer_changements(df_chg, erreurs)
        except Exception as e:
            messages.error(request, f"Erreur de lecture du fichier Changements : {e}")

    # Construction du message récapitulatif par ligne
    resume = []
    if f_dpt: 
        resume.append(f"Départements: {crees_dpts} ligne(s) importée(s) avec succès")
    if f_unite: 
        resume.append(f"Unités: {crees_unites} ligne(s) importée(s) avec succès")
    if f_collab: 
        resume.append(f"Collaborateurs: {crees_collabs} ligne(s) importée(s) avec succès")
    if f_chg: 
        resume.append(f"Changements d'affectation: {crees_chgs} ligne(s) importée(s) avec succès")

    messages.success(request, "Importation terminée : " + " | ".join(resume))
    
    if erreurs:
        messages.warning(request, f"{len(erreurs)} avertissement(s) : " + " | ".join(erreurs[:5]))

    return render(request, "import_data/import.html", {
        "form": MultipleImportForm(),
        "template_de_base": template_de_base,
    })

# ============================================================
# Récupère les collaborateurs "réels" (hors départs) avec leur RU d'affichage
# (RU d'accueil si changement, sinon RU habituel)
# ============================================================
def get_collaborateurs_reels(departements):
    collaborateurs = (
        Collaborateur.objects
        .filter(departement_id__in=departements)
        .select_related("departement", "unite", "ru_it")
    )

    ids = list(collaborateurs.values_list("it", flat=True))

    # Dernière déclaration pertinente (C, D, A, V) par collaborateur
    declarations = (
        declaration_effectif.objects
        .filter(collaborateur_it_id__in=ids, nature__in=["C", "D", "A", "V"])
        .select_related("nv_Ru")
        .order_by("collaborateur_it_id", "-date", "-id")
    )

    dernier_par_collab = {}
    for d in declarations:
        cid = d.collaborateur_it_id
        if cid not in dernier_par_collab:
            dernier_par_collab[cid] = d

    resultats = []
    for c in collaborateurs:
        d = dernier_par_collab.get(c.it)

        # Exclusion des départs
        if d and d.nature == "D":
            continue

        if d and d.nature == "C" and d.nv_Ru_id:
            ru_affiche = d.nv_Ru_id
        else:
            ru_affiche = c.ru_it_id

        resultats.append({
            "collaborateur": c,
            "ru": ru_affiche,
        })

    return resultats

@role_required(["HRBP", "DRH", "ADMIN"])
def export_effectif_reel(request):
    it = request.session.get("it")
    role = request.session.get("role")

    if role == "HRBP":
        departements_qs = Departement.objects.filter(HRBP_id=it)
    elif role == "DRH":
        departements_qs = Departement.objects.filter(DRH_id=it)
    elif role == "ADMIN":
        departements_qs = Departement.objects.filter(ADMIN_id=it)
    else:
        departements_qs = Departement.objects.none()

    departements = list(departements_qs.values_list("abreviation", flat=True))

    resultats = get_collaborateurs_reels(departements)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Effectif Réel"
    headers = ["Matricule", "IT", "Nom & Prénom", "Département", "Unité", "Lot", "RU(utilisateur)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in sorted(resultats, key=lambda x: (x["collaborateur"].departement_id or "", x["collaborateur"].nom_complete)):
        c = r["collaborateur"]
        ws.append([
            c.matricule,
            c.it,
            c.nom_complete,
            c.departement.abreviation if c.departement else "-",
            c.unite_id if c.unite_id else "-",
            c.lot,
            r["ru"] or "-",
        ])

    for col_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = max(length + 2, 12)

    today_str = timezone.localdate().strftime("%Y-%m-%d")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="effectif_reel_{today_str}.xlsx"'
    wb.save(response)

    return response